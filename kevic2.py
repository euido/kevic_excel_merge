import sys
import os
import json
import traceback
from PyQt5.QtWidgets import (
    QApplication, QVBoxLayout, QPushButton, QLabel, QFileDialog, QWidget, QLineEdit
)
from PyQt5.QtCore import QThread, pyqtSignal
from openpyxl import load_workbook
from openpyxl.styles import Alignment


CONFIG_FILE = "config.json"  # Config file to save paths


def parse_search_values(search_value):
    """Parsing the search range string into a list of integers."""
    values = set()
    for part in search_value.split(","):
        part = part.strip()
        if "-" in part:
            try:
                start, end = map(int, part.split("-"))
                values.update(range(start, end + 1))
            except ValueError:
                pass
        else:
            try:
                values.add(int(part))
            except ValueError:
                pass
    return list(values)


class ProcessThread(QThread):
    progress = pyqtSignal(str)
    finished = pyqtSignal(int)

    def __init__(self, file_a_path, search_values, file_b_path, save_path):
        super().__init__()
        self.file_a_path = file_a_path
        self.search_values = search_values
        self.file_b_path = file_b_path
        self.save_path = save_path

    def run(self):
        try:
            self.progress.emit("현황표,현품표 로딩중!!!! 조금만 기다려주세요!! 뿌잉...")
            file_a_workbook = load_workbook(self.file_a_path)
            file_a_sheet = file_a_workbook.active

            file_b_workbook = load_workbook(self.file_b_path)
            base_sheet = file_b_workbook.active

            self.progress.emit("Starting scan from B8...")
            match_count = 0

            for row in range(8, file_a_sheet.max_row + 1):
                b_value = file_a_sheet.cell(row=row, column=2).value  # B열
                d_value = file_a_sheet.cell(row=row, column=4).value  # D열
                e_value = file_a_sheet.cell(row=row, column=5).value  # E열
                i_value_d = file_a_sheet.cell(row=row, column=9).value  # I열
                i_value_f = file_a_sheet.cell(row=row, column=6).value  # F열
                m_value = file_a_sheet.cell(row=row, column=13).value  # M열

                if b_value in self.search_values:
                    self.progress.emit(f"Match found at row {row}")
                    match_count += 1

                    new_sheet = file_b_workbook.copy_worksheet(base_sheet)
                    new_sheet.title = f"{b_value}"

                    new_sheet['C6'] = b_value
                    new_sheet['D15'] = d_value
                    new_sheet['D17'] = e_value
                    new_sheet['D22'] = i_value_d
                    new_sheet['F22'] = i_value_d
                    new_sheet['D26'] = m_value

                    # Enable text wrapping in D26
                    new_sheet['C6'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    new_sheet['D15'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    new_sheet['D17'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    new_sheet['D22'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    new_sheet['F22'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    new_sheet['D26'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            new_file_path = os.path.join(self.save_path, f"현품표.xlsx")

            file_b_workbook.save(new_file_path)
            self.progress.emit(f"Saved to new file: {new_file_path}")
            self.finished.emit(match_count)

        except Exception as e:
            self.progress.emit(f"Error: {str(e)}")
            self.finished.emit(0)


class MainApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("KEVIC 현품표 Ver1.0.0")
        self.resize(500, 300)

        self.layout = QVBoxLayout()

        self.input_label = QLabel("Ex) 범위를 지정해주세요 (2,3,4 or 2-5):")
        self.search_input = QLineEdit()
        self.file_a_btn = QPushButton("현황표를 선택하세요")
        self.file_b_btn = QPushButton("현품표 서식을 선택하세요")
        self.scan_btn = QPushButton("현품표 생성")
        self.status_label = QLabel("엑셀 파일을 선택하세요")

        self.layout.addWidget(self.input_label)
        self.layout.addWidget(self.search_input)
        self.layout.addWidget(self.file_a_btn)
        self.layout.addWidget(self.file_b_btn)
        self.layout.addWidget(self.scan_btn)
        self.layout.addWidget(self.status_label)
        self.setLayout(self.layout)

        self.file_a_btn.clicked.connect(self.select_file_a)
        self.file_b_btn.clicked.connect(self.select_file_b)
        self.scan_btn.clicked.connect(self.scan_b_column)

        self.file_a_path = None
        self.file_b_path = None
        self.load_saved_paths()

    def save_paths(self):
        """Save paths to config.json for persistence."""
        config_data = {
            "file_a_path": self.file_a_path,
            "file_b_path": self.file_b_path,
        }
        with open(CONFIG_FILE, "w") as config_file:
            json.dump(config_data, config_file)

    def load_saved_paths(self):
        """Load saved paths from config.json if available."""
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, "r") as config_file:
                try:
                    config_data = json.load(config_file)
                    self.file_a_path = config_data.get("file_a_path", None)
                    self.file_b_path = config_data.get("file_b_path", None)
                    if self.file_a_path:
                        print(f"[DEBUG] Loaded File A: {self.file_a_path}")
                    if self.file_b_path:
                        print(f"[DEBUG] Loaded File B: {self.file_b_path}")
                except Exception as e:
                    print(f"[ERROR] Failed to load config: {e}")

    def select_file_a(self):
        path, _ = QFileDialog.getOpenFileName(self, "현황표")
        if path:
            self.file_a_path = path
            self.status_label.setText(f"현황표: {path}")
            self.save_paths()
            print(f"[DEBUG] Selected File A: {path}")

    def select_file_b(self):
        path, _ = QFileDialog.getOpenFileName(self, "현품표 서식")
        if path:
            self.file_b_path = path
            self.status_label.setText(f"현품표: {path}")
            self.save_paths()
            print(f"[DEBUG] Selected File B: {path}")

    def scan_b_column(self):
        if not self.file_a_path or not self.file_b_path:
            self.status_label.setText("현황표 및 현품표 Excel 파일을 모두 선택하세요")
            return

        search_value = self.search_input.text()
        if not search_value.strip():
            self.status_label.setText("범위를 지정해주세요!!!!.")
            return

        try:
            search_values = parse_search_values(search_value)
        except Exception as e:
            self.status_label.setText("Error!!!!!")
            return


        self.thread = ProcessThread(self.file_a_path, search_values, self.file_b_path, os.getcwd())
        self.thread.progress.connect(self.update_status)
        self.thread.finished.connect(self.on_finished)
        self.thread.start()

    def update_status(self, message):
        print(f"[DEBUG] Status: {message}")
        self.status_label.setText(message)

    def on_finished(self, count):
        if count > 0:
            self.status_label.setText(f"{count}개 sheets 완성.")
        else:
            self.status_label.setText("엑셀이 열려 있거나 숫자가 아닌 문자열 입니다!!! 확인 후 다시 시도해 주세요")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainApp()
    window.show()
    sys.exit(app.exec_())
