import sys
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from PyQt5.QtCore import QThread, pyqtSignal
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QLabel
)
#pyinstaller -w -F kevic_excel_merge.py
class FormatExcelThread(QThread):
    progress = pyqtSignal(str)

    def __init__(self, file_path):
        super().__init__()
        self.file_path = file_path

    def run(self):
        try:
            self.progress.emit("병합 작업 시작...")
            df = pd.read_excel(self.file_path, header=None)
            self.update_original_file(df)
            self.progress.emit("병합 및 저장 완료!")
        except Exception as e:
            self.progress.emit(f"오류 발생: {e}")

    def update_original_file(self, df):
        wb = load_workbook(self.file_path)
        ws = wb.active

        current_start_row = None
        current_value = None
        merged_b_values = []
        merged_c_values = []
        merged_d_values = []
        merged_e_values = []
        merged_f_values = []
        merged_g_values = []
        merged_h_values = []

        for index, row in df.iterrows():
            a_value = row[0]
            b_value = row[1]
            c_value = row[2]
            d_value = row[3]
            e_value = row[4]
            f_value = row[5]
            g_value = row[6]
            h_value = row[7]

            if pd.notna(a_value):
                if current_value is not None and current_start_row is not None:
                    ws.merge_cells(start_row=current_start_row, start_column=10, end_row=index, end_column=10)
                    ws.cell(row=current_start_row, column=10).value = current_value
                    ws.cell(row=current_start_row, column=10).alignment = Alignment(horizontal="center", vertical="center")

                    ws.merge_cells(start_row=current_start_row, start_column=11, end_row=index, end_column=11)
                    ws.cell(row=current_start_row, column=11).value = "\n".join(merged_b_values)
                    ws.cell(row=current_start_row, column=11).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    ws.merge_cells(start_row=current_start_row, start_column=12, end_row=index, end_column=12)
                    ws.cell(row=current_start_row, column=12).value = "\n".join(merged_c_values)
                    ws.cell(row=current_start_row, column=12).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    ws.merge_cells(start_row=current_start_row, start_column=13, end_row=index, end_column=13)
                    ws.cell(row=current_start_row, column=13).value = "\n".join(merged_d_values)
                    ws.cell(row=current_start_row, column=13).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    ws.merge_cells(start_row=current_start_row, start_column=14, end_row=index, end_column=14)
                    ws.cell(row=current_start_row, column=14).value = "\n".join(merged_e_values)
                    ws.cell(row=current_start_row, column=14).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    ws.merge_cells(start_row=current_start_row, start_column=15, end_row=index, end_column=15)
                    ws.cell(row=current_start_row, column=15).value = "\n".join(merged_f_values)
                    ws.cell(row=current_start_row, column=15).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    ws.merge_cells(start_row=current_start_row, start_column=16, end_row=index, end_column=16)
                    ws.cell(row=current_start_row, column=16).value = "\n".join(merged_g_values)
                    ws.cell(row=current_start_row, column=16).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    ws.merge_cells(start_row=current_start_row, start_column=8, end_row=index, end_column=8)
                    ws.cell(row=current_start_row, column=8).value = "\n".join(merged_h_values)
                    ws.cell(row=current_start_row, column=8).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    ws.merge_cells(start_row=current_start_row, start_column=17, end_row=index, end_column=17)
                    ws.cell(row=current_start_row, column=17).value = "\n".join(merged_h_values)
                    ws.cell(row=current_start_row, column=17).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                current_start_row = index + 1
                current_value = a_value
                merged_b_values = []
                merged_c_values = []
                merged_d_values = []
                merged_e_values = []
                merged_f_values = []
                merged_g_values = []
                merged_h_values = []

            if pd.notna(b_value):
                merged_b_values.append(str(b_value))
            if pd.notna(c_value):
                merged_c_values.append(str(c_value))
            if pd.notna(d_value):
                merged_d_values.append(str(d_value))
            if pd.notna(e_value):
                merged_e_values.append(str(e_value))
            if pd.notna(f_value):
                merged_f_values.append(str(f_value))
            if pd.notna(g_value):
                merged_g_values.append(str(g_value))
            if pd.notna(h_value):
                merged_h_values.append(str(h_value))

        if current_value is not None and current_start_row is not None:
            ws.merge_cells(start_row=current_start_row, start_column=10, end_row=len(df), end_column=10)
            ws.cell(row=current_start_row, column=10).value = current_value
            ws.cell(row=current_start_row, column=10).alignment = Alignment(horizontal="center", vertical="center")

        wb.save(self.file_path)

class CopyDataThread(QThread):
    progress = pyqtSignal(str)

    def __init__(self, source_file, target_file):
        super().__init__()
        self.source_file = source_file
        self.target_file = target_file

    def run(self):
        try:
            self.progress.emit("데이터 복사 작업 시작...")
            source_wb = load_workbook(self.source_file)
            source_ws = source_wb.active
            target_wb = load_workbook(self.target_file)
            target_ws = target_wb.active

            target_row_b = 8
            target_row_c = 8
            target_row_i = 8
            target_row_m = 8
            target_row_d = 8
            target_row_e = 8

            for row in range(2, source_ws.max_row + 1):
                value = source_ws.cell(row=row, column=10).value
                if value:
                    target_ws.cell(row=target_row_b, column=2).value = value
                    target_row_b += 1

                value = source_ws.cell(row=row, column=11).value
                if value:
                    target_ws.cell(row=target_row_c, column=3).value = value
                    target_row_c += 1

                value = source_ws.cell(row=row, column=17).value
                if value:
                    target_ws.cell(row=target_row_i, column=9).value = value
                    target_ws.cell(row=target_row_i, column=11).value = value
                    target_row_i += 1

                value = source_ws.cell(row=row, column=16).value
                if value:
                    target_ws.cell(row=target_row_m, column=13).value = value
                    target_row_m += 1

                value = source_ws.cell(row=row, column=12).value
                if value:
                    target_ws.cell(row=target_row_d, column=4).value = value
                    target_row_d += 1

                value = source_ws.cell(row=row, column=13).value
                if value:
                    target_ws.cell(row=target_row_e, column=5).value = value
                    target_row_e += 1

            target_wb.save(self.target_file)
            self.progress.emit("데이터 복사가 완료되었습니다!")
        except Exception as e:
            self.progress.emit(f"오류 발생: {e}")

class ExcelFormatterApp(QWidget):
    def __init__(self):
        super().__init__()
        self.file_path = None
        self.target_file_path = None
        self.initUI()

    def initUI(self):
        self.setWindowTitle('KEVIC 현황표 Ver1.0.0')
        self.setGeometry(300, 300, 400, 300)

        self.layout = QVBoxLayout()

        self.status_label = QLabel('엑셀 파일을 선택하세요.')
        self.layout.addWidget(self.status_label)

        self.file_button = QPushButton('1. 병합 엑셀 선택')
        self.file_button.clicked.connect(self.select_file)
        self.layout.addWidget(self.file_button)

        self.merge_button = QPushButton('2. 병합 및 저장 (원본 파일 수정)')
        self.merge_button.clicked.connect(self.start_format_excel_thread)
        self.layout.addWidget(self.merge_button)

        self.target_file_button = QPushButton('3. 자재실사 양식 선택')
        self.target_file_button.clicked.connect(self.select_target_file)
        self.layout.addWidget(self.target_file_button)

        self.copy_button = QPushButton('4. 자재실사 양식 데이터 저장')
        self.copy_button.clicked.connect(self.start_copy_data_thread)
        self.layout.addWidget(self.copy_button)

        self.result_label = QLabel('')
        self.layout.addWidget(self.result_label)

        self.setLayout(self.layout)

    def select_file(self):
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(self, "엑셀 파일 선택", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        if file_path:
            self.file_path = file_path
            self.status_label.setText(f"선택된 파일: {file_path}")

    def select_target_file(self):
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(self, "대상 엑셀 파일 선택", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        if file_path:
            self.target_file_path = file_path
            self.status_label.setText(f"대상 파일 선택: {file_path}")

    def start_format_excel_thread(self):
        if not self.file_path:
            self.result_label.setText("파일을 먼저 선택하세요!")
            return

        self.merge_button.setEnabled(False)
        self.copy_button.setEnabled(False)
        self.format_thread = FormatExcelThread(self.file_path)
        self.format_thread.progress.connect(self.update_status)
        self.format_thread.finished.connect(self.enable_buttons)
        self.format_thread.start()

    def start_copy_data_thread(self):
        if not self.file_path or not self.target_file_path:
            self.result_label.setText("원본 및 대상 파일을 모두 선택하세요!")
            return

        self.merge_button.setEnabled(False)
        self.copy_button.setEnabled(False)
        self.copy_thread = CopyDataThread(self.file_path, self.target_file_path)
        self.copy_thread.progress.connect(self.update_status)
        self.copy_thread.finished.connect(self.enable_buttons)
        self.copy_thread.start()

    def update_status(self, message):
        self.result_label.setText(message)

    def enable_buttons(self):
        self.merge_button.setEnabled(True)
        self.copy_button.setEnabled(True)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelFormatterApp()
    window.show()
    sys.exit(app.exec_())
